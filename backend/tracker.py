"""
tracker.py — Módulo de seguimiento de predicciones en Excel.
Guarda predicciones antes de cada partido y actualiza con resultados reales.
"""
import os
from datetime import datetime, date as date_type

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ─── Columnas ────────────────────────────────────────────────────────────────
HEADERS = [
    "Fecha", "Liga", "Local", "Visitante",
    "% Local", "% Empate", "% Visitante", "Predicción",
    "Resultado", "Ganador Real", "¿Correcto?", "Estado",
    "Guardado el", "Home ID", "Away ID",
]

C_DATE      = 1   # Fecha del partido
C_LEAGUE    = 2
C_HOME      = 3
C_AWAY      = 4
C_HOME_PCT  = 5
C_DRAW_PCT  = 6
C_AWAY_PCT  = 7
C_PREDICTED = 8   # "Local" | "Empate" | "Visitante"
C_RESULT    = 9   # "2–1"
C_WINNER    = 10  # ganador real
C_CORRECT   = 11  # "Sí" | "No"
C_STATUS    = 12  # "pendiente" | "jugado"
C_SAVED_AT  = 13
C_HOME_ID   = 14
C_AWAY_ID   = 15

# ─── Estilos ─────────────────────────────────────────────────────────────────
FILL_HEADER  = PatternFill("solid", fgColor="1A3A5C")
FILL_PENDING = PatternFill("solid", fgColor="FFF9C4")   # amarillo claro
FILL_CORRECT = PatternFill("solid", fgColor="C8E6C9")   # verde claro
FILL_WRONG   = PatternFill("solid", fgColor="FFCDD2")   # rojo claro

FONT_HEADER  = Font(bold=True, color="FFFFFF")
FONT_BOLD    = Font(bold=True)

COL_WIDTHS = [12, 22, 26, 26, 9, 9, 9, 12, 10, 14, 10, 11, 18, 9, 9]


class ExcelTracker:
    def __init__(self, path: str):
        self.path = path
        if os.path.exists(path):
            self.wb = openpyxl.load_workbook(path)
            self.ws = self.wb.active
        else:
            self.wb = openpyxl.Workbook()
            self.ws = self.wb.active
            self.ws.title = "Predicciones"
            self._init_sheet()

    # ─── Inicialización ──────────────────────────────────────────────────────

    def _init_sheet(self):
        for col, header in enumerate(HEADERS, 1):
            cell = self.ws.cell(row=1, column=col, value=header)
            cell.fill = FILL_HEADER
            cell.font = FONT_HEADER
            cell.alignment = Alignment(horizontal="center")
        for col, width in enumerate(COL_WIDTHS, 1):
            self.ws.column_dimensions[get_column_letter(col)].width = width
        self.ws.freeze_panes = "A2"  # congelar encabezado

    # ─── Consultas ───────────────────────────────────────────────────────────

    def already_tracked(self, home: str, away: str, match_date: str) -> bool:
        """¿Ya tenemos predicción para este partido?"""
        prefix = str(match_date)[:10]
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            row_date = str(row[C_DATE - 1])[:10]
            if row[C_HOME - 1] == home and row[C_AWAY - 1] == away and row_date == prefix:
                return True
        return False

    def get_pending_past_matches(self) -> list[dict]:
        """Partidos predichos cuya fecha ya pasó y aún no tienen resultado."""
        today = date_type.today()
        result = []
        for row_idx, row in enumerate(self.ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[C_STATUS - 1] != "pendiente":
                continue
            date_str = str(row[C_DATE - 1])[:10]
            try:
                match_date = datetime.strptime(date_str, "%Y-%m-%d").date()
                if match_date < today:
                    result.append({
                        "row":       row_idx,
                        "date":      date_str,
                        "league":    row[C_LEAGUE - 1],
                        "home":      row[C_HOME - 1],
                        "away":      row[C_AWAY - 1],
                        "predicted": row[C_PREDICTED - 1],
                        "home_id":   row[C_HOME_ID - 1],
                        "away_id":   row[C_AWAY_ID - 1],
                    })
            except (ValueError, TypeError):
                pass
        return result

    # ─── Escritura ───────────────────────────────────────────────────────────

    def add_prediction(self, match: dict, prediction: dict, league: str):
        """Agrega una nueva fila con la predicción."""
        probs       = prediction["probabilities"]
        winner_key  = max(probs, key=probs.get)
        winner_label = {"home_win": "Local", "draw": "Empate", "away_win": "Visitante"}[winner_key]

        row_data = [
            match.get("date", "")[:10],           # C_DATE
            league,                                 # C_LEAGUE
            match.get("home", ""),                  # C_HOME
            match.get("away", ""),                  # C_AWAY
            probs["home_win"],                      # C_HOME_PCT
            probs["draw"],                          # C_DRAW_PCT
            probs["away_win"],                      # C_AWAY_PCT
            winner_label,                           # C_PREDICTED
            "",                                     # C_RESULT    (vacío)
            "",                                     # C_WINNER    (vacío)
            "",                                     # C_CORRECT   (vacío)
            "pendiente",                            # C_STATUS
            datetime.now().strftime("%Y-%m-%d %H:%M"),  # C_SAVED_AT
            match.get("home_id", ""),               # C_HOME_ID
            match.get("away_id", ""),               # C_AWAY_ID
        ]

        next_row = self.ws.max_row + 1
        for col, val in enumerate(row_data, 1):
            cell = self.ws.cell(row=next_row, column=col, value=val)
            cell.fill = FILL_PENDING
            if col in (C_HOME_PCT, C_DRAW_PCT, C_AWAY_PCT):
                cell.number_format = "0.0"
        self.save()

    def update_result(self, row_idx: int, home_goals: int, away_goals: int,
                      winner: str, predicted: str):
        """Actualiza una fila con el resultado real y si la predicción fue correcta."""
        correct = "Sí" if winner == predicted else "No"
        fill    = FILL_CORRECT if correct == "Sí" else FILL_WRONG

        self.ws.cell(row=row_idx, column=C_RESULT,  value=f"{home_goals}–{away_goals}")
        self.ws.cell(row=row_idx, column=C_WINNER,  value=winner)
        self.ws.cell(row=row_idx, column=C_CORRECT, value=correct)
        self.ws.cell(row=row_idx, column=C_STATUS,  value="jugado")

        for col in range(1, len(HEADERS) + 1):
            self.ws.cell(row=row_idx, column=col).fill = fill

        self.save()

    def save(self):
        try:
            self.wb.save(self.path)
        except PermissionError:
            # El archivo puede estar abierto en Excel
            alt = self.path.replace(".xlsx", f"_auto_{datetime.now().strftime('%H%M%S')}.xlsx")
            self.wb.save(alt)
            print(f"⚠️  El Excel estaba abierto. Guardado como: {os.path.basename(alt)}")

    # ─── Estadísticas ────────────────────────────────────────────────────────

    def summary(self) -> dict:
        total = correct = wrong = pending = 0
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            status = row[C_STATUS - 1]
            if status == "jugado":
                total += 1
                if row[C_CORRECT - 1] == "Sí":
                    correct += 1
                else:
                    wrong += 1
            elif status == "pendiente":
                pending += 1
        return {
            "total":    total,
            "correct":  correct,
            "wrong":    wrong,
            "pending":  pending,
            "accuracy": round(correct / total * 100, 1) if total else 0.0,
        }
